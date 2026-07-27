"""Gera sugestões de de-para entre as entidades do CNMP (dim_unidade) e as
unidades da planilha da SAP-SP, para curadoria manual.

Não há chave comum: o CNMP usa nomes honoríficos em caixa alta, muitas vezes
sem a cidade; a SAP usa tipo abreviado + cidade, às vezes agregando unidades
("Penit. X + PC de Andradina"). O matching é por interseção de tokens
normalizados (sem acentos/pontuação, abreviações expandidas, stopwords
removidas), usando o nome SAP + município como alvo.

Imprime o dict DEPARA no formato de src/modulos/sap/depara_unidades.py, com o
nome CNMP como comentário. Casos ambíguos ou de score baixo saem marcados com
"REVISAR" (e os de score muito baixo saem comentados) — revise-os manualmente
antes de colar no arquivo curado. Como no gerador de aliases, o script não
sobrescreve o arquivo curado para não perder edições manuais.

Execute (do diretório python/):
    uv run python scripts/gerar_depara_unidades_sap.py
"""

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

_PASTA_DOWNLOADS = Path(__file__).resolve().parent.parent / "downloads"
_PLANILHA_SAP = _PASTA_DOWNLOADS / "sap/sap_nova_corrigida.xlsx"
_INSPECAO_CNMP = _PASTA_DOWNLOADS / "cnmp/inspecao_api_resolucao_277.json"

# Só o ambiente 462 (unidades comuns): a SAP-SP não administra as unidades
# militares do ambiente 282.
_AMBIENTE = "462"
_FORMULARIOS = ("1322", "1342")

_ABREVIACOES = {
    "penit": "penitenciaria",
    "cdp": "centro detencao provisoria",
    "cpp": "centro progressao penitenciaria",
    "cr": "centro ressocializacao",
    "pc": "prisao civil",
    "fem": "feminino",
    "rsa": "",
    "prsa": "",
    "adp": "",
    "app": "",
}

_STOPWORDS = {"de", "da", "do", "dos", "das", "e", "a", "o", "regime",
              "fechado", "semiaberto", "aberto", "i", "ii", "iii", "iv"}

_SCORE_MINIMO = 0.7
_MARGEM_UNICO = 0.15


def _tokens(texto: str) -> set[str]:
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    texto = texto.lower().replace("_", " ")
    texto = re.sub(r"[^a-z0-9 ]", " ", texto)
    resultado: set[str] = set()
    for token in texto.split():
        for sub in _ABREVIACOES.get(token, token).split():
            if sub and sub not in _STOPWORDS:
                resultado.add(sub)
    return resultado


def _score(tokens_cnmp: set[str], tokens_sap: set[str]) -> float:
    if not tokens_cnmp or not tokens_sap:
        return 0.0
    intersecao = tokens_cnmp & tokens_sap
    return len(intersecao) / min(len(tokens_cnmp), len(tokens_sap))


def _entidades_cnmp() -> dict[int, str]:
    data = json.loads(_INSPECAO_CNMP.read_text(encoding="utf-8"))
    entidades: dict[int, str] = {}
    for formulario_id in _FORMULARIOS:
        for item in data["ambientes"][_AMBIENTE]["formularios"][formulario_id]["entidades"]:
            for entidade_id, detalhe in item.items():
                entidades[int(entidade_id)] = detalhe["descricao"]
    return entidades


def _unidades_sap() -> list[tuple[str, str]]:
    workbook = openpyxl.load_workbook(_PLANILHA_SAP, read_only=True)
    planilha = workbook[workbook.sheetnames[0]]
    linhas = list(planilha.iter_rows(min_row=2, values_only=True))
    return [(linha[1], linha[2] or "") for linha in linhas if linha[1]]


def main() -> None:
    entidades = _entidades_cnmp()
    unidades = _unidades_sap()
    tokens_sap = [(nome, _tokens(f"{nome} {municipio}")) for nome, municipio in unidades]

    contagem = Counter()
    print("DEPARA: dict[int, str] = {")
    for entidade_id, nome_cnmp in sorted(entidades.items()):
        tokens_cnmp = _tokens(nome_cnmp)
        ranking = sorted(
            ((_score(tokens_cnmp, t), nome) for nome, t in tokens_sap), reverse=True
        )
        melhor_score, melhor_nome = ranking[0]
        segundo_score = ranking[1][0]

        comentario = " ".join(nome_cnmp.split())
        if melhor_score >= _SCORE_MINIMO and melhor_score - segundo_score >= _MARGEM_UNICO:
            contagem["automatico"] += 1
            print(f"    {entidade_id}: {melhor_nome!r},  # {comentario}")
        elif melhor_score >= _SCORE_MINIMO:
            contagem["revisar_ambiguo"] += 1
            print(
                f"    {entidade_id}: {melhor_nome!r},"
                f"  # REVISAR ambíguo ({melhor_score:.2f} vs {segundo_score:.2f},"
                f" 2º: {ranking[1][1]!r}) — {comentario}"
            )
        else:
            contagem["revisar_baixo"] += 1
            print(
                f"    # {entidade_id}: {melhor_nome!r},"
                f"  # REVISAR score baixo ({melhor_score:.2f}) — {comentario}"
            )
    print("}")
    print()
    print(f"# automáticos: {contagem['automatico']}, ambíguos p/ revisar: "
          f"{contagem['revisar_ambiguo']}, score baixo (comentados): {contagem['revisar_baixo']}")


if __name__ == "__main__":
    main()
