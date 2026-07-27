"""Camada silver do módulo SAP: lê a planilha de unidades prisionais da SAP-SP
no Lakehouse mp_bronze (Files/sap/), aplica as transformações puras de
transform.py e recarrega a tabela sap_unidade no Warehouse mp_silver.

A planilha é atualizada manualmente (upload pelo portal do Fabric para
Files/sap/); esta carga apenas a materializa como tabela.

Execute:
    uv run python -m src.modulos.sap.etl.load_silver
"""

import io
import logging

import openpyxl
from sqlalchemy import Engine, text

from src.infra.lakehouse import download_bytes
from src.infra.warehouse import get_silver_engine
from src.modulos.cnmp.etl.load_silver import recarregar_tabela
from src.modulos.sap.transform import COLUNAS_SAP_UNIDADE, linhas_sap_unidade

logging.basicConfig(level=logging.WARNING)
logging.getLogger("src.modulos.sap.etl").setLevel(logging.INFO)
logger = logging.getLogger(__name__)

CAMINHO_PLANILHA = "sap/sap_nova_corrigida.xlsx"

# Fabric Warehouse (esta edição) rejeita PRIMARY KEY e DEFAULT no CREATE TABLE
# (ver load_silver do cnmp); a chave natural é unidade_nome, garantida única
# na transformação.
DDL_SAP = """
IF OBJECT_ID('sap_unidade', 'U') IS NULL
CREATE TABLE sap_unidade (
    ano                     INT NULL,
    unidade_nome            VARCHAR(200) NOT NULL,
    municipio               VARCHAR(120) NULL,
    regional                VARCHAR(120) NULL,
    cod_ibge                INT NULL,
    custodia                VARCHAR(50) NULL,
    comarca                 VARCHAR(150) NULL,
    raj                     VARCHAR(120) NULL,
    destinacao              VARCHAR(200) NULL,
    faccoes                 VARCHAR(MAX) NULL,
    servidores              INT NULL,
    policiais_penais        INT NULL,
    presos                  INT NULL,
    capacidade              INT NULL,
    superlotacao_percentual DECIMAL(9, 2) NULL,
    mortes_naturais         INT NULL,
    mortes_intervencao      INT NULL,
    mortes_indeterminadas   INT NULL,
    providencias            VARCHAR(MAX) NULL,
    deecrim                 VARCHAR(50) NULL,
    medicos                 INT NULL,
    dentistas               INT NULL,
    psicologos              INT NULL,
    assistentes_sociais     INT NULL,
    psiquiatras             INT NULL,
    presos_estudam          INT NULL,
    presos_trabalham        INT NULL,
    genero                  VARCHAR(30) NULL
);
"""


def criar_schema(engine: Engine) -> None:
    with engine.begin() as conn:
        conn.execute(text(DDL_SAP.strip().rstrip(";")))
    logger.info("Schema sap_unidade criado/verificado")


def ler_planilha(conteudo: bytes) -> list[dict]:
    """Abre o xlsx (em bytes) e devolve as linhas transformadas."""
    workbook = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True)
    planilha = workbook[workbook.sheetnames[0]]
    linhas = list(planilha.iter_rows(values_only=True))
    return linhas_sap_unidade(list(linhas[0]), linhas[1:])


def carregar_silver(engine: Engine, caminho_planilha: str = CAMINHO_PLANILHA) -> None:
    criar_schema(engine)
    conteudo = download_bytes(caminho_planilha)
    linhas = ler_planilha(conteudo)
    logger.info("planilha %s: %d unidades", caminho_planilha, len(linhas))
    recarregar_tabela(engine, "sap_unidade", COLUNAS_SAP_UNIDADE, linhas)


if __name__ == "__main__":
    carregar_silver(get_silver_engine())
